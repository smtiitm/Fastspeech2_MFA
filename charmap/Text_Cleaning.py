#!/usr/bin/env python
# coding: utf-8

# In[2]:


############################################################
#Author : Bhagyashree
#Date : 1st Sept, 2020
#Purpose : Text Cleaning
#Input : Text file after timestamp removal
#Output : Text file after cleaning data
############################################################


# In[3]:


import nltk
import numpy
import xlrd
import openpyxl 
import re
import sys

# In[21]:

#file1 = open("recent_deliverables_dec2020/Corporate_Law/Hindi/ankita objects 02_Hindi_new.txt","r+",encoding='utf-8') 
file1 = open(sys.argv[1],"r+",encoding='utf-8') 
data = file1.read()
#print(data)
file1.close()


# In[22]:

wb_obj = openpyxl.load_workbook(sys.argv[2]) 
sheet_obj = wb_obj.active 

#data = re.sub('[A-Z]*', '',data)
#print(data)
data = data.replace('?','')
data = data.replace('  ',' ')
data = data.replace(';','')
data = data.replace(')','')
data = data.replace('(','')
data = data.replace('!','')
data = data.replace(' – ',' ')
data = data.replace('-',' ')
data = data.replace('।','')
data = data.replace('&','')
data = data.replace('’','')
data = data.replace('‘','')
data = data.replace(':','')
data = data.replace(',','')
data = data.replace('/','')
data = data.replace(',','')
data = data.replace('.','')
data = data.replace('|','')
m_row = sheet_obj.max_row 
line = data

for i in range(1,m_row+1):
    num = sheet_obj.cell(row = i, column = 1).value 
    word = sheet_obj.cell(row = i, column = 2).value
    #print(num)
    #print(word)
    line = line.replace(str(num), word)
#print(line)
#' '.join(line.split())
#print(line) 
file1 = open(sys.argv[3],"w+",encoding='utf-8') 
#file1 = open("recent_deliverables_dec2020/Corporate_Law/Hindi/ankita objects 02_Hindi.txt","w+",encoding='utf-8') 
file1.write(line)
file1.close()

